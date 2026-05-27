"""Provenance .xlsx sidecar — one row per number on the deck.

Walks `canonical_store` for the ticker, plus the memo computed values
and the macro snapshot, and emits an Excel workbook the analyst can
open alongside the .pptx to audit every figure: source provider,
URL, the year/quarter the value belongs to, and when it was fetched.

This file is the answer to "where did each number come from, and for
which period?" — the single sidecar the user asked for. LLM-generated
prose cells are marked `LLM (Gemini)` with the upstream numeric anchor
ID list rather than left ambiguous.

Layout:
  Sheet "Provenance"
    Slide | Section | Metric | Value | Source | Source URL | Data Period | Fetched At | Notes
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from src.services.canonical_store import (
    get_all_fields, get_observations_by_provider,
)


# ── Field → slide / section / human label mapping ────────────────
#
# Maps canonical_store field names to where they appear on the deck.
# Fields not in the map are still surfaced (as "Other / Other") so the
# analyst sees the full inventory; this only controls grouping.
_FIELD_MAP: dict[str, tuple[str, str, str]] = {
    # field name              (slide, section, human label)
    "company_profile":        ("Slide 1", "Header", "Company profile"),
    "quote":                  ("Slide 1", "Header", "Quote"),
    "rating_split":           ("Slide 1", "Analyst Consensus", "Rating split (buy/hold/sell)"),
    "consensus_target":       ("Slide 1", "Analyst Consensus", "Average target price"),
    "valuation_forward":      ("Slide 1", "Key Data / Slide 2 Table",
                                "Forward P/E + Q+1 forecasts"),
    "valuation_historical":   ("Slide 3", "P/E Chart", "Historical P/E series"),
    "historical_prices":      ("Slide 3", "52-Week Price Chart",
                                "Daily closes + 52w range + performance buckets"),
    "income_statement_quarterly": ("Slide 3", "Earnings History Chart",
                                     "Per-quarter actual vs estimate surprise track"),
    "income_statement_annual":    ("Slide 2", "Estimates Table (history)",
                                     "Annual revenue / EBITDA / NI history"),
    "broker_actions":         ("Slide 2", "Catalysts (track-record anchor)", "Recent broker actions"),
    "dividends_payments":     ("Slide 1", "Key Data", "Dividend yield"),
    "earnings_calendar":      ("Slide 1", "Header", "Next earnings date"),
}


# ── Helpers ──────────────────────────────────────────────────────

def _value_repr(v: Any) -> str:
    """Compact text rendering of a canonical_value for the Value column.
    Dicts and lists are collapsed to their key list so the row stays
    legible — the analyst can drill into the raw .json fixtures if a
    deeper view is needed."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        if abs(v) >= 1e9:
            return f"{v:,.2f}"
        if abs(v) >= 1e6:
            return f"{v:,.0f}"
        if abs(v) >= 1:
            return f"{v:,.2f}"
        return f"{v:.4f}"
    if isinstance(v, dict):
        keys = list(v.keys())[:6]
        more = " …" if len(v) > 6 else ""
        return f"<dict: {', '.join(keys)}>{more}"
    if isinstance(v, list):
        return f"<list[{len(v)}]>"
    s = str(v)
    return s if len(s) <= 80 else s[:77] + "…"


def _source_url(provider: str, ticker: str) -> str:
    """Best-effort URL for the canonical source. We don't keep per-cell
    URLs in canonical_store (the raw_observations table holds those),
    so this gives the analyst the right page to land on for a given
    provider."""
    p = (provider or "").lower()
    if "investing" in p:
        return "https://www.investing.com (per-equity slug, see data/investing/)"
    if "marketscreener" in p:
        return "https://www.marketscreener.com"
    if "yahoo" in p or "yfinance" in p:
        return f"https://finance.yahoo.com/quote/{ticker}"
    if "bloomberg" in p:
        return "Analyst Bloomberg export (data/bloomberg/)"
    if p in ("imf", "imf weo"):
        return "https://www.imf.org/external/datamapper"
    if p in ("wb", "world bank", "worldbank"):
        return "https://api.worldbank.org/v2"
    if "gemini" in p or "llm" in p:
        return "https://aistudio.google.com (Gemini API)"
    return ""


def _data_period_from_value(field: str, value: Any) -> str:
    """Pull a human period tag from the value payload when present.
    Different providers store period tags under different keys —
    this is the de-facto union over what we've seen in the wild."""
    if not isinstance(value, dict):
        return ""
    # Macro fields: explicit year tag with source label.
    if field == "company_profile":
        bits = []
        if value.get("macro_year"):
            bits.append(f"macro {value.get('macro_year')}")
        if value.get("gdp_growth_fcst_year"):
            bits.append(f"IMF GDP {value.get('gdp_growth_fcst_year')}F")
        if value.get("inflation_fcst_year"):
            bits.append(f"IMF infl {value.get('inflation_fcst_year')}F")
        if bits:
            return ", ".join(bits)
    # Forecast bundle: surface fy1_year / next_q_period as the period anchor.
    if field == "valuation_forward":
        bits = []
        if value.get("next_q_period"):
            bits.append(str(value["next_q_period"]))
        if value.get("fy1_year"):
            bits.append(f"FY{value['fy1_year']}E")
        return ", ".join(bits)
    # Generic keys some providers stamp.
    for k in ("period", "as_of", "year", "fiscal_year", "report_year", "asof"):
        if value.get(k):
            return str(value[k])
    return ""


# ── Main entry ───────────────────────────────────────────────────

def _explode_dict_value(slide: str, section: str, label_prefix: str,
                          d: dict, src: str, src_url: str,
                          fetched_at: str, notes: str) -> list[list[str]]:
    """Expand a dict-valued canonical_store cell into one row per scalar
    sub-key, so the analyst sees every number on its own line instead of
    a `<dict: ...>` placeholder. Skips nested dicts/lists (the parent row
    still emits as a summary)."""
    out_rows = []
    SKIP_KEYS = {"summary"}   # long-form prose, not a numeric provenance row
    for k, v in d.items():
        if k in SKIP_KEYS:
            continue
        if v is None:
            continue
        if isinstance(v, (dict, list)):
            # Skip nested structures from the explosion; the parent row
            # surfaces them with a `<dict>` / `<list>` placeholder so the
            # analyst knows to inspect them in raw fixtures.
            continue
        # Format numbers compactly; leave strings as-is.
        if isinstance(v, (int, float)):
            disp = (f"{v:,.4f}" if abs(v) < 1 else
                    f"{v:,.2f}" if abs(v) < 1e6 else
                    f"{v:,.0f}")
        else:
            s = str(v)
            disp = s if len(s) <= 80 else s[:77] + "…"
        out_rows.append([
            slide, section, f"{label_prefix} · {k}",
            disp, src, src_url, "", fetched_at, notes,
        ])
    return out_rows


def write_provenance_xlsx(ticker: str, out_path: Path,
                            memo_data: Optional[dict] = None,
                            payload: Any = None,
                            peer_rows: Optional[list[dict]] = None,
                            ) -> Optional[Path]:
    """Walk canonical_store + memo + macro for `ticker` and write an
    .xlsx audit sheet to `out_path`. Returns the path on success, None
    if openpyxl is unavailable (kept optional so the deck still ships).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Provenance"
    headers = ["Slide", "Section", "Metric", "Value", "Source",
                 "Source URL", "Data Period", "Fetched At", "Notes"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D1117")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    rows: list[list[str]] = []

    # 1. Every canonical_store cell for the ticker — parent summary row
    #    plus per-sub-key explosion for dict-valued cells so every number
    #    on the deck is individually traceable.
    cv = get_all_fields(ticker)
    for field, c in sorted(cv.items()):
        slide, section, label = _FIELD_MAP.get(field, ("Other", "Other", field))
        period = _data_period_from_value(field, c.value)
        src = (c.canonical_source or "").strip()
        src_url = _source_url(c.canonical_source, ticker)
        fetched_at = c.last_refreshed_at.strftime("%Y-%m-%d %H:%M UTC")
        notes = (c.notes or "").strip()
        rows.append([
            slide, section, label,
            _value_repr(c.value), src, src_url,
            period, fetched_at, notes,
        ])
        # Dict explosion — one row per scalar sub-key. Keeps the analyst
        # from having to open raw JSON to see e.g. `current_price`,
        # `market_cap`, `mean_target_price`, `fwd_pe`, individually.
        if isinstance(c.value, dict):
            rows.extend(_explode_dict_value(slide, section, label,
                                              c.value, src, src_url,
                                              fetched_at, notes))
        # Historical price series: surface 52w stats (high / low / last /
        # period return) since the slide shows them and the full daily
        # series is too noisy to print row-by-row.
        if field == "historical_prices" and isinstance(c.value, list) and c.value:
            try:
                vals = [float(pt["close"]) for pt in c.value
                        if isinstance(pt, dict) and isinstance(pt.get("close"), (int, float))]
                if len(vals) >= 2:
                    hi = max(vals); lo = min(vals)
                    first = vals[0]; last = vals[-1]
                    ret_pct = (last - first) / first * 100.0 if first else None
                    stats = [
                        ("52w high",  f"{hi:,.4f}" if hi < 1 else f"{hi:,.2f}"),
                        ("52w low",   f"{lo:,.4f}" if lo < 1 else f"{lo:,.2f}"),
                        ("Last close", f"{last:,.4f}" if last < 1 else f"{last:,.2f}"),
                        ("# observations", str(len(vals))),
                    ]
                    if ret_pct is not None:
                        stats.append(("52w return %", f"{ret_pct:+.2f}%"))
                    for k, v in stats:
                        rows.append([slide, section, f"{label} · {k}",
                                       v, src, src_url, "", fetched_at, ""])
            except (KeyError, ValueError, TypeError):
                pass
        # Quarterly surprise track: one row per (period, eps_surprise_pct)
        if (field == "income_statement_quarterly" and isinstance(c.value, dict)
            and isinstance(c.value.get("surprise_history"), list)):
            for r in (c.value["surprise_history"] or [])[:8]:
                if not isinstance(r, dict): continue
                p = r.get("period") or r.get("date") or ""
                sp = r.get("eps_surprise_pct")
                ra = r.get("revenue_actual")
                re_ = r.get("revenue_estimate")
                if isinstance(sp, (int, float)) and p:
                    rows.append([slide, section, f"{label} · {p} EPS surprise",
                                   f"{sp:+.2f}%", src, src_url, str(p),
                                   fetched_at, ""])
                if isinstance(ra, (int, float)) and isinstance(re_, (int, float)) and p:
                    rows.append([slide, section, f"{label} · {p} Revenue act/est",
                                   f"{ra:,.0f} / {re_:,.0f}", src, src_url, str(p),
                                   fetched_at, ""])

    # 2. Memo-derived fields — every Q+1 consensus that fed slide 2 plus
    #    derived YoY / QoQ deltas. Full coverage: revenue, EPS, EBITDA,
    #    Net Income, and the implied upside on slide 1.
    if memo_data:
        nq_src = memo_data.get("next_quarter_consensus_source")
        nq_label = memo_data.get("next_quarter_label") or memo_data.get("next_earnings_label") or ""
        cascade_note = "Cascade: MS quarterly → Investing → Yahoo → MS annual÷4"
        memo_keys_quarterly = [
            ("next_quarter_consensus_revenue", "Q+1 consensus Revenue"),
            ("next_quarter_consensus_eps",     "Q+1 consensus EPS"),
            ("next_quarter_consensus_ebitda",  "Q+1 consensus EBITDA"),
            ("next_quarter_consensus_ni",      "Q+1 consensus Net Income"),
        ]
        for key, human in memo_keys_quarterly:
            v = memo_data.get(key)
            if v is not None:
                rows.append([
                    "Slide 2", "Estimates Table",
                    f"{human} ({nq_label})" if nq_label else human,
                    _value_repr(v),
                    nq_src or "MarketScreener",
                    _source_url(nq_src or "MarketScreener", ticker),
                    nq_label, "", cascade_note,
                ])
        # Implied upside / target reflected in the deck.
        if memo_data.get("implied_upside_pct") is not None:
            rows.append([
                "Slide 1", "Analyst Consensus", "Implied upside %",
                _value_repr(memo_data["implied_upside_pct"]),
                "Derived",
                "(target_mean / quote.price - 1) × 100",
                "", "", "Per-cell formula",
            ])
        # Next-quarter YoY/QoQ derivations (when memo computed them).
        # _compute_memo uses *_growth_pct suffix for revenue/NI; the
        # thesis renderer additionally computes per-metric YoY/QoQ but
        # doesn't write back. Both shapes covered here.
        yoy_qoq_keys = [
            ("next_quarter_yoy_revenue_growth_pct", "Q+1 YoY revenue %"),
            ("next_quarter_yoy_ni_growth_pct",       "Q+1 YoY net income %"),
            ("next_quarter_yoy_revenue",             "Q+1 YoY revenue %"),
            ("next_quarter_yoy_eps",                 "Q+1 YoY EPS %"),
            ("next_quarter_yoy_ebitda",              "Q+1 YoY EBITDA %"),
            ("next_quarter_yoy_ni",                  "Q+1 YoY net income %"),
            ("next_quarter_qoq_revenue",             "Q+1 QoQ revenue %"),
            ("next_quarter_qoq_eps",                 "Q+1 QoQ EPS %"),
            ("next_quarter_qoq_ebitda",              "Q+1 QoQ EBITDA %"),
            ("next_quarter_qoq_ni",                  "Q+1 QoQ net income %"),
        ]
        for key, human in yoy_qoq_keys:
            v = memo_data.get(key)
            if isinstance(v, (int, float)):
                rows.append([
                    "Slide 2", "Estimates Table", human,
                    f"{v:+.2f}%", "Derived",
                    "consensus vs same-Q prior-year (YoY) or prior-Q (QoQ)",
                    nq_label, "", "Computed in _compute_memo / render time",
                ])

    # 3. Peer comparables — emit one row per peer × per shown metric.
    #    Slide 3 shows: COMPANY | TICKER | MCAP | P/E | P/TBV | DIV YIELD |
    #    1Y RETURN. Trace each cell so the analyst can drill into source.
    if peer_rows:
        for pr in peer_rows:
            if not isinstance(pr, dict): continue
            ptick = pr.get("ticker") or pr.get("symbol") or ""
            pname = pr.get("name") or pr.get("company") or ptick
            metrics_to_emit = [
                ("market_cap", "Market cap"),
                ("forward_pe", "Forward P/E"),
                ("price_to_book", "P/TBV"),
                ("price_to_tangible_book", "P/TBV"),
                ("dividend_yield", "Dividend yield %"),
                ("return_1y_pct", "1Y return %"),
            ]
            for src_key, human in metrics_to_emit:
                v = pr.get(src_key)
                if v is None: continue
                rows.append([
                    "Slide 3", "Peer Comparables",
                    f"{pname} ({ptick}) · {human}",
                    _value_repr(v),
                    (pr.get("source") or "yfinance"),
                    _source_url(pr.get("source") or "yfinance", ptick),
                    "", "", "",
                ])

    # 4. Company-disclosed quarterly figures (Phase 1 of the disclosed-
    #    source pipeline). When `data/disclosed/{ticker}.json` exists,
    #    each disclosed quarter is emitted as one row per metric so the
    #    audit trail shows the exact IR-portal PDF the value came from.
    try:
        from src.services.disclosed_loader import load_disclosed
        d_payload = load_disclosed(ticker)
    except Exception:
        d_payload = None
    if d_payload:
        company_name = d_payload.get("company") or ticker
        sd_map = d_payload.get("_source_documents") or {}
        for q in d_payload.get("quarterly", []) or []:
            if not isinstance(q, dict): continue
            period = str(q.get("period") or "").strip()
            if not period: continue
            doc = q.get("source_doc") or ""
            url = sd_map.get(period) or ""
            metric_pairs = [
                ("operating_income", "Operating Income (raw currency)"),
                ("net_interest_income", "Net Interest Income"),
                ("fee_income_net", "Fee & Commission Income (net)"),
                ("net_income", "Net Income"),
                ("eps", "EPS"),
            ]
            for key, human in metric_pairs:
                v = q.get(key)
                if not isinstance(v, (int, float)): continue
                rows.append([
                    "Slide 3", "Earnings History (Disclosed)",
                    f"{human} · {period}",
                    _value_repr(v),
                    f"{company_name} IR",
                    url, period, "",
                    f"Standalone-quarter value from {doc}",
                ])

    # 5. MS annual forecasts — emit FY year × metric rows when present.
    #    These back the annual-fallback table on slide 2 and the forward
    #    P/E bars on slide 3.
    annual = None
    if payload is not None:
        msa = getattr(payload, "ms_annual_forecasts", None)
        if isinstance(msa, dict):
            annual = msa.get("annual") if isinstance(msa.get("annual"), dict) else msa
    if annual and isinstance(annual.get("periods"), list):
        periods = annual["periods"]
        for arr_key, human in [("net_sales", "Net sales"),
                                  ("ebitda",    "EBITDA"),
                                  ("net_income", "Net income")]:
            arr = annual.get(arr_key) or []
            for i, p in enumerate(periods):
                if i >= len(arr): continue
                v = arr[i]
                if not isinstance(v, (int, float)): continue
                rows.append([
                    "Slide 2", "Annual Forecasts (MS)",
                    f"{human} · {p}",
                    _value_repr(v),
                    "MarketScreener",
                    "https://www.marketscreener.com",
                    str(p), "",
                    "Annual estimate; per-quarter breakdown not published",
                ])

    # 3. LLM-generated prose (thesis, catalysts, risks, watch_list, highlights).
    # These are not in canonical_store; we surface them explicitly so the
    # analyst sees what the model produced AND that no numerics live here.
    try:
        from src.services.llm_summary import generate_summary
        llm = generate_summary(ticker)
    except Exception:
        llm = None
    if llm:
        as_of = (llm.get("as_of") or "").replace("T", " ")[:19]
        ctx_hash = llm.get("context_hash") or ""
        thesis = (llm.get("thesis_paragraph") or "").strip()
        if thesis:
            rows.append([
                "Slide 2", "Investment Thesis", "Executive summary (4-sentence)",
                thesis[:120] + ("…" if len(thesis) > 120 else ""),
                "LLM (Gemini)",
                _source_url("gemini", ticker),
                "", as_of, f"context_hash={ctx_hash}; numeric-trace validated",
            ])
        for i, c in enumerate(llm.get("catalysts") or [], start=1):
            rows.append([
                "Slide 2", "Catalysts", f"Catalyst #{i}",
                str(c)[:120],
                "LLM (Gemini)",
                _source_url("gemini", ticker),
                "", as_of, f"context_hash={ctx_hash}",
            ])
        for i, c in enumerate(llm.get("risks") or [], start=1):
            rows.append([
                "Slide 2", "Risks", f"Risk #{i}",
                str(c)[:120],
                "LLM (Gemini)",
                _source_url("gemini", ticker),
                "", as_of, f"context_hash={ctx_hash}",
            ])
        for i, c in enumerate(llm.get("watch_list") or [], start=1):
            rows.append([
                "Slide 2", "What to Watch", f"Watch #{i}",
                str(c)[:120],
                "LLM (Gemini)",
                _source_url("gemini", ticker),
                "", as_of, f"context_hash={ctx_hash}",
            ])
        for it in (llm.get("highlights") or []):
            cat = (it.get("category") or "").strip().upper()
            body = (it.get("body") or "").strip()
            if cat and body:
                rows.append([
                    "Slide 1", f"Analyst Highlights — {cat}", "Pill body",
                    body[:120],
                    "LLM (Gemini)",
                    _source_url("gemini", ticker),
                    "", as_of, f"context_hash={ctx_hash}",
                ])

    # Sort by slide then section for a clean read.
    def _slide_sort(r):
        s = (r[0] or "")
        # "Slide 1" < "Slide 2" < ... < "Other"
        if s.startswith("Slide "):
            try:
                return (0, int(s.split(" ", 1)[1]), r[1])
            except (ValueError, IndexError):
                pass
        return (1, 0, r[1])
    rows.sort(key=_slide_sort)
    for r in rows:
        ws.append(r)

    # Column widths sized to typical content. openpyxl doesn't auto-fit.
    widths = [9, 32, 36, 32, 16, 48, 22, 22, 50]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"

    # Metadata sheet — quick at-a-glance: ticker + run timestamp + counts.
    meta = wb.create_sheet("About")
    meta.append(["Ticker", ticker])
    meta.append(["Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["Rows", len(rows)])
    meta.append([])
    meta.append(["Notes", (
        "Every numeric value on the deck should trace to a row here. "
        "LLM-generated prose has no numeric source — its numbers are "
        "validated against the canonical_store anchors at generation time."
    )])
    for col in ("A", "B"):
        meta.column_dimensions[col].width = 28
    meta["A1"].font = Font(bold=True)
    meta["A2"].font = Font(bold=True)
    meta["A3"].font = Font(bold=True)
    meta["A5"].font = Font(bold=True)
    meta["B5"].alignment = Alignment(wrap_text=True, vertical="top")
    meta.row_dimensions[5].height = 60

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
