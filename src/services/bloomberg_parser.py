"""
Bloomberg manual-export parser.

Parses the two xlsx layouts we actually have:

  data/bloomberg/<YAHOO_TICKER>_cons_q.xlsx  — quarterly consensus grid
  data/bloomberg/<YAHOO_TICKER>_FA.xlsx      — multi-year financial analysis

Not compatible with earnings-preview.v2's bloomberg_parser.py (which expects
the Bloomberg MODL "Single Period" sheet with per-broker detail). This
parser is for the cons_q + FA format exported from Bloomberg Terminal
(our actual inputs).

Files are dropped into data/bloomberg/ by hand (committed to git). The
loader is a no-op when files are absent — pipeline falls through to
MarketScreener/Yahoo unchanged.

Key ideas:
  * Parse by label + period-header row, never by fixed row/column indexes
    (cons_q rows vary per ticker: ADNOCGAS has 18 line items, SAFCO has 19
    due to an extra DPS row).
  * Normalize Bloomberg labels and field codes to the metric keys used
    elsewhere in the pipeline (revenue, ebitda, ebit, net_income, eps, …).
  * Return None on missing files, not empty objects — integration code
    can branch on `bundle is not None`.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from src.config import root

log = logging.getLogger(__name__)


# ── Label / field-code normalization ─────────────────────────────────────

# Consensus-quarterly line-item labels → canonical metric key.
# Match is substring-based (lowercased) so variants like "EPS, Adj+" and
# "EPS Adjusted" both hit the same key.
_CONSQ_LABEL_MAP: list[tuple[str, str]] = [
    ("eps, adj", "eps_adj"),
    ("eps adj", "eps_adj"),
    ("eps, gaap", "eps_gaap"),
    ("eps gaap", "eps_gaap"),
    ("revenue", "revenue"),
    ("gross margin", "gross_margin_pct"),
    ("operating income (ebit)", "ebit"),
    ("ebitda", "ebitda"),
    ("pre-tax profit", "pretax_profit"),
    ("pre tax profit", "pretax_profit"),
    ("net income, adj", "net_income_adj"),
    ("net income adj", "net_income_adj"),
    ("net income, gaap", "net_income_gaap"),
    ("net income gaap", "net_income_gaap"),
    ("net debt", "net_debt"),
    ("bps", "bps"),
    ("cps", "cps"),
    ("dps", "dps"),
    ("return on equity", "roe_pct"),
    ("return on assets", "roa_pct"),
    ("depreciation", "depreciation"),
    ("free cash flow", "fcf"),
    ("capex", "capex"),
    ("net asset value", "nav"),
]

# FA field-code (col B) → canonical key. Falls back to label if code is blank.
_FA_CODE_MAP: dict[str, str] = {
    "HISTORICAL_MARKET_CAP": "market_cap",
    "CUR_MKT_CAP": "market_cap",
    "CASH_AND_MARKETABLE_SECURITIES": "cash",
    "PFD_EQTY_MINORTY_INTEREST": "preferred_minority",
    "SHORT_AND_LONG_TERM_DEBT": "total_debt",
    "ENTERPRISE_VALUE": "enterprise_value",
    "SALES_REV_TURN": "revenue",
    "SALES_GROWTH": "revenue_growth_pct",
    "EBITDA": "ebitda_or_margin",  # shared code; resolved by label
    "GROSS_PROFIT": "gross_profit_or_margin",
    "EARN_FOR_COMMON": "net_income_or_margin",
    "IS_DIL_EPS_CONT_OPS": "eps",
    "DILUTED_EPS_AFT_XO_ITEMS_GROWTH": "eps_growth_pct",
    "CF_CASH_FROM_OPER": "cfo",
    "CAPITAL_EXPEND": "capex",
    "CF_FREE_CASH_FLOW": "fcf",
}

# FA label suffixes that mark a "margin" row rather than the raw value row.
_FA_MARGIN_MARKERS = ("margin %", "margin%", "margin")


# ── Dataclasses ──────────────────────────────────────────────────────────

@dataclass
class BloombergConsensusQuarter:
    period_label: str          # e.g. "Q1 2026", "Q4 2025"
    period_end: str            # ISO YYYY-MM-DD
    is_estimate: bool
    currency: str
    metrics: dict[str, tuple[float | None, int | None]] = field(default_factory=dict)
    # metric_key -> (mean_value, n_analysts). n_analysts is None for the
    # actual column and for rows with no `#` column populated.


@dataclass
class BloombergAnnualPeriod:
    period_label: str          # "FY 2021", "FY 2026 Est", "Current/LTM"
    period_end: str            # ISO YYYY-MM-DD
    is_estimate: bool
    is_ltm: bool
    currency: str
    metrics: dict[str, float | None] = field(default_factory=dict)


@dataclass
class BloombergBundle:
    ticker: str                                      # Yahoo-format, e.g. "ADNOCGAS.AE"
    bbg_ticker: str = ""                             # e.g. "ADNOCGAS UH"
    company_name: str = ""
    currency: str = ""
    consensus_quarterly: list[BloombergConsensusQuarter] = field(default_factory=list)
    annuals: list[BloombergAnnualPeriod] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def latest_actual_quarter(self) -> BloombergConsensusQuarter | None:
        return next((q for q in self.consensus_quarterly if not q.is_estimate), None)

    def next_estimate_quarter(self) -> BloombergConsensusQuarter | None:
        return next((q for q in self.consensus_quarterly if q.is_estimate), None)


# ── Parsing helpers ──────────────────────────────────────────────────────

_QUARTER_RE = re.compile(r"Q([1-4])\s+(\d{4})", re.I)
_FY_RE = re.compile(r"FY\s*(\d{4})", re.I)
_EST_RE = re.compile(r"\bEst\b", re.I)


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _as_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in {"—", "-", "N/A", "N/M", "NM"}:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _as_int(v) -> int | None:
    f = _as_float(v)
    if f is None:
        return None
    try:
        return int(f)
    except (TypeError, ValueError):
        return None


def _iso_date(v) -> str:
    """Convert MM/DD/YYYY, datetime, or ISO string to YYYY-MM-DD. Empty on fail."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            pass
    s = _norm(v)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        mm, dd, yyyy = m.groups()
        return f"{yyyy}-{int(mm):02d}-{int(dd):02d}"
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s)
    if m:
        y, mo, d = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return ""


def _match_consq_key(label: str) -> str | None:
    lab = label.lower().strip()
    for needle, key in _CONSQ_LABEL_MAP:
        if needle in lab:
            return key
    return None


# ── Public loader ────────────────────────────────────────────────────────

def bloomberg_dir() -> Path:
    return root() / "data" / "bloomberg"


def _ticker_files(ticker: str) -> tuple[Path, Path]:
    d = bloomberg_dir()
    return d / f"{ticker}_cons_q.xlsx", d / f"{ticker}_FA.xlsx"


def bloomberg_coverage(ticker: str) -> dict:
    """Return {cons_q: bool, fa: bool} for quick UI/coverage checks."""
    cq, fa = _ticker_files(ticker)
    return {"cons_q": cq.is_file(), "fa": fa.is_file()}


def load_bloomberg_bundle(ticker: str) -> BloombergBundle | None:
    """Load both xlsx files for a Yahoo ticker. Return None if neither exists.

    Missing-but-parse-error cases return a bundle with warnings so the
    caller can surface what went wrong without aborting the pipeline.
    """
    cq_path, fa_path = _ticker_files(ticker)
    if not cq_path.is_file() and not fa_path.is_file():
        return None

    bundle = BloombergBundle(ticker=ticker)

    if cq_path.is_file():
        try:
            _parse_cons_q(cq_path, bundle)
        except Exception as exc:  # noqa: BLE001 — surface as a warning, not a hard fail
            log.warning("Bloomberg cons_q parse failed for %s: %s", ticker, exc)
            bundle.warnings.append(f"cons_q parse failed: {exc}")

    if fa_path.is_file():
        try:
            _parse_fa(fa_path, bundle)
        except Exception as exc:  # noqa: BLE001
            log.warning("Bloomberg FA parse failed for %s: %s", ticker, exc)
            bundle.warnings.append(f"FA parse failed: {exc}")

    return bundle


# ── cons_q parser ────────────────────────────────────────────────────────

def _parse_cons_q(path: Path, bundle: BloombergBundle) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    header_row: int | None = None   # row with "Q4 2025 Act" / "Q1 2026 Est"
    date_row: int | None = None     # row with "3 Months Ending" + `#` markers

    # Scan the first 15 rows for the header band and capture metadata from col A.
    max_scan = min(ws.max_row or 0, 15)
    for r in range(1, max_scan + 1):
        a = _norm(ws.cell(row=r, column=1).value)
        b = _norm(ws.cell(row=r, column=2).value)
        if not a and not b:
            continue
        a_low = a.lower()

        if r == 1 and a and not bundle.bbg_ticker:
            # "ADNOCGAS UH Equity" → "ADNOCGAS UH"
            bundle.bbg_ticker = re.sub(r"\s+Equity\s*$", "", a, flags=re.I).strip()

        if a_low == "currency" and b:
            bundle.currency = b.upper()
        elif a_low == "3 months ending":
            date_row = r
        elif b and _QUARTER_RE.search(b):
            header_row = r

    if header_row is None or date_row is None:
        bundle.warnings.append("cons_q: could not find header / date rows")
        return

    # Build (col, period_label, is_estimate) list from the header row, plus
    # the corresponding # column when present one column to the right.
    periods: list[tuple[int, int | None, str, bool, str]] = []
    # (value_col, analyst_count_col_or_None, period_label, is_estimate, period_end_iso)

    max_col = ws.max_column or 10
    c = 2
    while c <= max_col:
        val_label = _norm(ws.cell(row=header_row, column=c).value)
        if val_label:
            q_match = _QUARTER_RE.search(val_label)
            if q_match:
                is_est = bool(_EST_RE.search(val_label))
                period_label = f"Q{q_match.group(1)} {q_match.group(2)}"
                # The "#" count column sits immediately to the right of each est col.
                next_cell = _norm(ws.cell(row=date_row, column=c + 1).value)
                count_col = (c + 1) if next_cell == "#" else None
                # Period-end date from the date_row
                period_end = _iso_date(ws.cell(row=date_row, column=c).value)
                periods.append((c, count_col, period_label, is_est, period_end))
                c = (count_col + 1) if count_col else (c + 1)
                continue
        c += 1

    if not periods:
        bundle.warnings.append("cons_q: no period columns recognized")
        return

    # Build consensus quarter objects keyed by period, then walk rows to fill
    # metrics.
    quarters = [
        BloombergConsensusQuarter(
            period_label=pl,
            period_end=pe,
            is_estimate=is_est,
            currency=bundle.currency,
        )
        for (_vc, _cc, pl, is_est, pe) in periods
    ]

    # Walk from after the date_row to end
    for r in range(date_row + 1, (ws.max_row or date_row) + 1):
        label = _norm(ws.cell(row=r, column=1).value)
        if not label:
            continue
        key = _match_consq_key(label)
        if key is None:
            continue
        for idx, (val_col, count_col, *_rest) in enumerate(periods):
            raw_val = ws.cell(row=r, column=val_col).value
            raw_cnt = ws.cell(row=r, column=count_col).value if count_col else None
            quarters[idx].metrics[key] = (_as_float(raw_val), _as_int(raw_cnt))

    bundle.consensus_quarterly = quarters


# ── FA parser ────────────────────────────────────────────────────────────

def _parse_fa(path: Path, bundle: BloombergBundle) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    # Row 2 col 1: "Adnoc Gas PLC (ADNOCGAS UH) - BBG Adj Highlights"
    title = _norm(ws.cell(row=2, column=1).value)
    if title:
        m = re.match(r"^(.*?)\s*\(([^)]+)\)", title)
        if m:
            if not bundle.company_name:
                bundle.company_name = m.group(1).strip()
            if not bundle.bbg_ticker:
                bundle.bbg_ticker = m.group(2).strip()

    # Row 4: period-label row. Col 1 has "In Millions of <CCY>".
    header_row = 4
    date_row = 5
    a4 = _norm(ws.cell(row=header_row, column=1).value)
    m_ccy = re.search(r"In Millions of\s+([A-Z]{3})", a4, re.I)
    if m_ccy and not bundle.currency:
        bundle.currency = m_ccy.group(1).upper()

    # Discover period columns from row 4 (any cell that looks like "FY YYYY"
    # or "Current/LTM").
    max_col = ws.max_column or 12
    period_cols: list[tuple[int, str, bool, bool, str]] = []
    # (col, period_label, is_estimate, is_ltm, period_end_iso)

    for c in range(2, max_col + 1):
        lbl = _norm(ws.cell(row=header_row, column=c).value)
        if not lbl:
            continue
        is_ltm = ("LTM" in lbl.upper()) or ("CURRENT" in lbl.upper())
        is_est = bool(_EST_RE.search(lbl))
        fy_m = _FY_RE.search(lbl)
        if not (fy_m or is_ltm):
            continue
        period_end = _iso_date(ws.cell(row=date_row, column=c).value)
        period_cols.append((c, lbl, is_est, is_ltm, period_end))

    if not period_cols:
        bundle.warnings.append("FA: no period columns recognized")
        return

    annuals = [
        BloombergAnnualPeriod(
            period_label=pl,
            period_end=pe,
            is_estimate=is_est,
            is_ltm=is_ltm,
            currency=bundle.currency,
        )
        for (_c, pl, is_est, is_ltm, pe) in period_cols
    ]

    # Walk rows from header_row+2 downward.
    end_row = ws.max_row or header_row
    for r in range(header_row + 2, end_row + 1):
        label = _norm(ws.cell(row=r, column=1).value)
        code = _norm(ws.cell(row=r, column=2).value)
        if not label:
            continue
        if label.lower().startswith("source:"):
            break

        key = _fa_key_for(label, code)
        if key is None:
            continue

        for idx, (col, *_rest) in enumerate(period_cols):
            val = _as_float(ws.cell(row=r, column=col).value)
            if val is not None or key not in annuals[idx].metrics:
                annuals[idx].metrics[key] = val

    bundle.annuals = annuals


def _fa_key_for(label: str, code: str) -> str | None:
    """Resolve a label+code row to a canonical metric key."""
    lab_low = label.lower().strip()
    is_margin = any(m in lab_low for m in _FA_MARGIN_MARKERS)
    is_growth = "growth" in lab_low

    code = (code or "").upper()
    base = _FA_CODE_MAP.get(code)

    # Label-first fallback for rows where col-B code is blank.
    if base is None:
        if "revenue" in lab_low and not is_growth:
            return "revenue"
        if "gross profit" in lab_low and not is_margin:
            return "gross_profit"
        if "ebitda" in lab_low and not is_margin:
            return "ebitda"
        if "net income" in lab_low and not is_margin and not is_growth:
            return "net_income"
        if "eps" in lab_low and not is_growth:
            return "eps"
        if "cash from op" in lab_low:
            return "cfo"
        if "capital expend" in lab_low or "capex" in lab_low:
            return "capex"
        if "free cash flow" in lab_low:
            return "fcf"
        if "market cap" in lab_low:
            return "market_cap"
        if "enterprise value" in lab_low:
            return "enterprise_value"
        return None

    # Resolve shared codes (EBITDA, GROSS_PROFIT, EARN_FOR_COMMON) by the
    # margin/growth suffix in the label.
    if base == "ebitda_or_margin":
        return "ebitda_margin_pct" if is_margin else "ebitda"
    if base == "gross_profit_or_margin":
        return "gross_margin_pct" if is_margin else "gross_profit"
    if base == "net_income_or_margin":
        return "net_margin_pct" if is_margin else "net_income"
    return base


# ── Coverage helper (for scripts/UI) ─────────────────────────────────────

def list_covered_tickers(tickers: Iterable[str]) -> list[dict]:
    """Return a list of {ticker, cons_q, fa} for a set of tickers."""
    out: list[dict] = []
    for t in tickers:
        c = bloomberg_coverage(t)
        out.append({"ticker": t, **c})
    return out
